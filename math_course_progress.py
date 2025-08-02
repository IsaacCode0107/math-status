import pandas as pd
from datetime import datetime

# 課程資料結構
course_data = {
    "第三冊 第一章 三角函數": [
        {"code": "1-1_主題1", "name": "弧度量(A+B版)", "time": "40:39", "status": "已完成", "note": "基礎概念"},
        {"code": "1-1_主題2", "name": "扇形弧長與面積(A+B版)", "time": "15:20", "status": "已完成", "note": "應用計算"},
        {"code": "1-1_主題3", "name": "三角函數的圖形(A+B版)", "time": "01:12:32", "status": "已完成", "note": "圖形分析"},
        {"code": "1-1_主題4", "name": "三角方程式(A+B版)", "time": "23:55", "status": "進行中", "note": "方程求解"},
        {"code": "1-1_主題5", "name": "三角不等式(A+B版)", "time": "14:26", "status": "未開始", "note": "不等式解法"},
        {"code": "1-1_主題6", "name": "三角函數的極值(A+B版)", "time": "03:37", "status": "未開始", "note": "極值問題"},
        {"code": "1-1_主題7", "name": "三角函數的應用(A+B版)", "time": "13:39", "status": "未開始", "note": "實際應用"},
        {"code": "1-2_主題1", "name": "和差角公式(A版)", "time": "01:19:38", "status": "未開始", "note": "公式推導"},
        {"code": "1-2_主題2", "name": "平方差(課外補充)(A版)", "time": "06:08", "status": "未開始", "note": "補充內容"},
        {"code": "1-2_主題3", "name": "正切的和差角(A版)", "time": "27:06", "status": "未開始", "note": "正切公式"},
        {"code": "1-2_主題4", "name": "二倍角(A版)", "time": "28:12", "status": "未開始", "note": "倍角公式"},
        {"code": "1-2_主題5", "name": "二倍角(A版)", "time": "21:51", "status": "未開始", "note": "延續內容"},
        {"code": "1-2_主題6", "name": "半角公式(A版)", "time": "26:03", "status": "未開始", "note": "半角推導"},
        {"code": "1-3_主題1", "name": "正餘弦的疊合(A版)", "time": "40:16", "status": "未開始", "note": "函數疊合"},
        {"code": "1-3_主題2", "name": "正餘弦疊合的極值(A版)", "time": "44:27", "status": "未開始", "note": "極值計算"},
        {"code": "1-3_主題3", "name": "正餘弦疊合的應用問題(A版)", "time": "28:05", "status": "未開始", "note": "應用題型"}
    ],
    "第三冊 第二章 指數與對數函數": [
        {"code": "2-1_主題1", "name": "指數函數的圖形(A+B版)", "time": "42:24", "status": "未開始", "note": "圖形性質"},
        {"code": "2-1_主題2", "name": "指數比大小(A+B版)", "time": "09:01", "status": "未開始", "note": "大小比較"},
        {"code": "2-1_主題3", "name": "指數方程式與不等式(A+B版)", "time": "09:01", "status": "未開始", "note": "方程不等式"},
        {"code": "2-1_主題4", "name": "指數函數的極值(A+B版)", "time": "11:03", "status": "未開始", "note": "極值問題"},
        {"code": "2-1_主題5", "name": "指數函數的應用(A+B版)", "time": "08:15", "status": "未開始", "note": "實際應用"},
        {"code": "2-2_主題1", "name": "對數的定義(A+B版)", "time": "14:32", "status": "未開始", "note": "基本定義"},
        {"code": "2-2_主題2", "name": "對數律(A+B版)", "time": "01:04:23", "status": "未開始", "note": "對數運算律"},
        {"code": "2-2_主題3", "name": "對數方程(A+B版)", "time": "21:12", "status": "未開始", "note": "方程求解"},
        {"code": "2-2_主題4", "name": "科學記號與常用對數(A+B版)", "time": "24:07", "status": "未開始", "note": "科學應用"},
        {"code": "2-3_主題1", "name": "對數函數的圖形(A+B版)", "time": "44:59", "status": "未開始", "note": "圖形分析"},
        {"code": "2-3_主題2", "name": "對數比大小(A+B版)", "time": "12:06", "status": "未開始", "note": "大小比較"},
        {"code": "2-3_主題3", "name": "對數不等式(A+B版)", "time": "23:56", "status": "未開始", "note": "不等式解法"},
        {"code": "2-3_主題4", "name": "對數求極值(A+B版)", "time": "07:08", "status": "未開始", "note": "極值計算"},
        {"code": "2-3_主題5", "name": "對數應用問題(A+B版)", "time": "33:57", "status": "已完成", "note": "應用題型"}
    ],
    "第三冊 第三章 平面向量": [
        {"code": "3-1_主題1", "name": "平面向量的幾何表示法(A+B版)", "time": "05:53", "status": "未開始", "note": "幾何表示"},
        {"code": "3-1_主題2", "name": "平面向量的座標表示法(A+B版)", "time": "14:51", "status": "未開始", "note": "座標表示"},
        {"code": "3-1_主題3", "name": "向量的運算(A+B版)", "time": "24:45", "status": "已完成", "note": "基本運算"},
        {"code": "3-1_主題4", "name": "向量的線性組合(A+B版)", "time": "28:52", "status": "未開始", "note": "線性組合"},
        {"code": "3-1_主題5", "name": "共線理論(A+B版)", "time": "35:00", "status": "未開始", "note": "共線性質"},
        {"code": "3-1_主題6", "name": "面積比(A+B版)", "time": "07:25", "status": "未開始", "note": "面積計算"},
        {"code": "3-1_主題7", "name": "斜座標(A+B版)", "time": "19:00", "status": "未開始", "note": "座標系統"},
        {"code": "3-2_主題1", "name": "內積(A+B版)", "time": "26:00", "status": "未開始", "note": "內積運算"},
        {"code": "3-2_主題2", "name": "內積的性質(A+B版)", "time": "14:19", "status": "未開始", "note": "性質探討"},
        {"code": "3-2_主題3", "name": "正射影(A+B版)", "time": "26:11", "status": "未開始", "note": "投影計算"},
        {"code": "3-2_主題4", "name": "直線參數式(A+B版)", "time": "27:25", "status": "未開始", "note": "參數方程"},
        {"code": "3-2_主題5", "name": "直線的夾角(A+B版)", "time": "10:36", "status": "未開始", "note": "角度計算"},
        {"code": "3-2_主題6", "name": "柯西不等式(A版)", "time": "26:59", "status": "未開始", "note": "不等式理論"},
        {"code": "3-3_主題1", "name": "向量的面積公式(A版)", "time": "18:09", "status": "未開始", "note": "面積公式"},
        {"code": "3-3_主題2", "name": "二階行列式(A版)", "time": "29:15", "status": "未開始", "note": "行列式計算"},
        {"code": "3-3_主題3", "name": "克拉瑪公式(A版)", "time": "12:17", "status": "未開始", "note": "線性方程組"},
        {"code": "3-3_主題4", "name": "二元一次方程組的向量觀點(A版)", "time": "12:02", "status": "未開始", "note": "向量觀點"}
    ],
    "第四冊 第一章 空間向量": [
        {"code": "1-1_主題1", "name": "空間概念(A+B版)", "time": "23:07", "status": "未開始", "note": "基礎概念"},
        {"code": "1-1_主題2", "name": "二面角(A+B版)", "time": "30:00", "status": "未開始", "note": "角度計算"},
        {"code": "1-1_主題3", "name": "立體圖形計算(A+B版)", "time": "14:50", "status": "未開始", "note": "立體幾何"},
        {"code": "1-2_主題1", "name": "空間坐標系(A+B版)", "time": "32:39", "status": "未開始", "note": "坐標系統"},
        {"code": "1-2_主題2", "name": "空間向量的表示法(A+B版)", "time": "25:52", "status": "未開始", "note": "向量表示"},
        {"code": "1-2_主題3", "name": "空間向量的線性組合(A+B版)", "time": "11:01", "status": "未開始", "note": "線性運算"},
        {"code": "1-3_主題1", "name": "空間向量的內積(A版)", "time": "21:25", "status": "未開始", "note": "內積計算"},
        {"code": "1-3_主題2", "name": "柯西不等式(A版)", "time": "21:58", "status": "未開始", "note": "不等式"},
        {"code": "1-4_主題1", "name": "外積(A版)", "time": "18:27", "status": "未開始", "note": "外積運算"},
        {"code": "補充", "name": "內積與外積比較", "time": "06:34", "status": "未開始", "note": "概念比較"},
        {"code": "1-4_主題2", "name": "空間向量求體積(A版)", "time": "04:57", "status": "未開始", "note": "體積計算"},
        {"code": "1-4_主題3", "name": "三階行列式(A版)", "time": "24:50", "status": "未開始", "note": "行列式"},
        {"code": "1-4_主題4", "name": "三階行列式的應用(A版)", "time": "12:44", "status": "未開始", "note": "應用計算"},
        {"code": "數B補充", "name": "如何判斷圓錐截痕", "time": "04:27", "status": "未開始", "note": "幾何判斷"},
        {"code": "數B補充", "name": "球面與經緯度", "time": "20:27", "status": "未開始", "note": "球面幾何"}
    ],
    "第四冊第二章 空間的平面與直線(A版)": [
        {"code": "2-1_主題1", "name": "平面方程式", "time": "24:59", "status": "未開始", "note": "方程建立"},
        {"code": "2-1_主題2", "name": "平面的截距式", "time": "18:42", "status": "未開始", "note": "截距表示"},
        {"code": "2-1_主題3", "name": "平面的夾角", "time": "09:30", "status": "未開始", "note": "角度計算"},
        {"code": "2-1_主題4", "name": "點到平面的距離", "time": "24:21", "status": "未開始", "note": "距離公式"},
        {"code": "2-2_主題1", "name": "直線方程式", "time": "37:02", "status": "未開始", "note": "方程建立"},
        {"code": "2-2_主題2", "name": "直線與平面", "time": "29:32", "status": "未開始", "note": "關係分析"},
        {"code": "2-2_主題3", "name": "點與直線", "time": "06:50", "status": "未開始", "note": "位置關係"},
        {"code": "2-2_主題4", "name": "直線與直線", "time": "39:55", "status": "未開始", "note": "線線關係"}
    ],
    "第四冊 第三章 條件機率(A+B版)": [
        {"code": "3-1~3-2_主題1", "name": "條件機率", "time": "23:37", "status": "未開始", "note": "機率計算"},
        {"code": "3-2_主題2", "name": "獨立事件", "time": "30:05", "status": "未開始", "note": "事件獨立性"},
        {"code": "3-3_主題1", "name": "貝氏定理", "time": "27:36", "status": "未開始", "note": "定理應用"}
    ],
    "第四冊 第四章 矩陣": [
        {"code": "4-1_主題1", "name": "高斯消去法(A版)", "time": "37:19", "status": "未開始", "note": "解方程組"},
        {"code": "4-2_主題1", "name": "矩陣的加減法(A+B版)", "time": "11:50", "status": "未開始", "note": "基本運算"},
        {"code": "4-2_主題2", "name": "矩陣的乘法(A+B版)", "time": "21:19", "status": "未開始", "note": "乘法運算"},
        {"code": "4-2_主題3", "name": "矩陣的高次方(A+B版)", "time": "24:19", "status": "未開始", "note": "冪次計算"},
        {"code": "4-2_主題4", "name": "反方陣(A+B版)", "time": "46:26", "status": "未開始", "note": "逆矩陣"},
        {"code": "4-3_主題1", "name": "轉移矩陣(A版)", "time": "38:11", "status": "未開始", "note": "狀態轉移"},
        {"code": "4-3_主題2~3", "name": "平面線性變換(A版)", "time": "10:19", "status": "未開始", "note": "幾何變換"},
        {"code": "4-3_主題4~5", "name": "伸縮與推移矩陣(A版)", "time": "13:58", "status": "未開始", "note": "變換矩陣"},
        {"code": "4-3_主題6", "name": "鏡射矩陣(A版)", "time": "12:17", "status": "未開始", "note": "對稱變換"},
        {"code": "4-3_主題7", "name": "旋轉矩陣(A版)", "time": "24:45", "status": "未開始", "note": "旋轉變換"}
    ]
}

# 準備 Excel 資料
excel_data = []

for chapter_name, courses in course_data.items():
    # 添加章節標題行
    excel_data.append({
        '章節': chapter_name,
        '課程編號': '',
        '課程名稱': '',
        '時長': '',
        '狀態': '',
        '備註': '',
        '完成日期': ''
    })
    
    # 添加課程資料
    for course in courses:
        excel_data.append({
            '章節': '',
            '課程編號': course['code'],
            '課程名稱': course['name'],
            '時長': course['time'],
            '狀態': course['status'],
            '備註': course['note'],
            '完成日期': ''
        })

# 創建 DataFrame
df = pd.DataFrame(excel_data)

# 創建 Excel writer 物件
with pd.ExcelWriter('數學課程進度追蹤表.xlsx', engine='openpyxl') as writer:
    # 寫入主要工作表
    df.to_excel(writer, sheet_name='課程進度', index=False)
    
    # 獲取工作簿和工作表
    workbook = writer.book
    worksheet = writer.sheets['課程進度']
    
    # 設置欄寬
    worksheet.column_dimensions['A'].width = 25  # 章節
    worksheet.column_dimensions['B'].width = 15  # 課程編號
    worksheet.column_dimensions['C'].width = 35  # 課程名稱
    worksheet.column_dimensions['D'].width = 10  # 時長
    worksheet.column_dimensions['E'].width = 12  # 狀態
    worksheet.column_dimensions['F'].width = 20  # 備註
    worksheet.column_dimensions['G'].width = 12  # 完成日期
    
    # 創建樣式
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # 標題樣式
    header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # 章節標題樣式
    chapter_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    chapter_font = Font(color="FFFFFF", bold=True)
    
    # 狀態顏色
    completed_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    progress_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    not_started_fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # 邊框樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 設置標題行樣式
    for col in range(1, 8):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # 設置狀態下拉選單
    status_validation = DataValidation(
        type="list",
        formula1='"未開始,進行中,已完成"',
        allow_blank=False
    )
    status_validation.error = "請選擇：未開始、進行中、已完成"
    status_validation.errorTitle = "無效的狀態"
    worksheet.add_data_validation(status_validation)
    status_validation.add(f'E2:E{len(excel_data)+1}')
    
    # 設置資料樣式
    for row_idx, row_data in enumerate(excel_data, start=2):
        # 如果是章節標題行
        if row_data['課程編號'] == '':
            for col in range(1, 8):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.fill = chapter_fill
                cell.font = chapter_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        else:
            # 課程資料行
            for col in range(1, 8):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')
                
                # 狀態欄特殊處理
                if col == 5:  # 狀態欄
                    status = cell.value
                    if status == '已完成':
                        cell.fill = completed_fill
                    elif status == '進行中':
                        cell.fill = progress_fill
                    elif status == '未開始':
                        cell.fill = not_started_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # 時長欄置中
                elif col == 4:  # 時長欄
                    cell.alignment = Alignment(horizontal='center')
    
    # 創建統計工作表
    stats_data = [
        ['統計項目', '數值', '公式'],
        ['總課程數', f'=COUNTA(課程進度.B:B)-1', '計算課程編號欄位非空值'],
        ['已完成課程', f'=COUNTIF(課程進度.E:E,"已完成")', '計算狀態為"已完成"的課程'],
        ['進行中課程', f'=COUNTIF(課程進度.E:E,"進行中")', '計算狀態為"進行中"的課程'],
        ['未開始課程', f'=COUNTIF(課程進度.E:E,"未開始")', '計算狀態為"未開始"的課程'],
        ['完成率', f'=ROUND(C3/C2*100,1)&"%"', '已完成課程數/總課程數*100'],
        ['', '', ''],
        ['狀態說明', '', ''],
        ['已完成', '課程已學習完畢', ''],
        ['進行中', '課程正在學習中', ''],
        ['未開始', '課程尚未開始學習', ''],
        ['', '', ''],
        ['使用說明', '', ''],
        ['1. 直接修改"狀態"欄位', '使用下拉選單選擇', ''],
        ['2. 填寫"完成日期"', '格式：YYYY/MM/DD', ''],
        ['3. 更新"備註"', '記錄學習心得或進度', ''],
        ['4. 統計會自動計算', '無需手動更新', '']
    ]
    
    stats_df = pd.DataFrame(stats_data[1:], columns=stats_data[0])
    stats_df.to_excel(writer, sheet_name='統計與說明', index=False)
    
    # 設置統計工作表樣式
    stats_worksheet = writer.sheets['統計與說明']
    stats_worksheet.column_dimensions['A'].width = 20
    stats_worksheet.column_dimensions['B'].width = 25
    stats_worksheet.column_dimensions['C'].width = 30
    
    # 統計工作表標題樣式
    for col in range(1, 4):
        cell = stats_worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

print("Excel 檔案已生成：數學課程進度追蹤表.xlsx")
print("\n功能說明：")
print("1. 📊 主要工作表「課程進度」- 包含所有課程資料")
print("2. 📈 統計工作表「統計與說明」- 自動統計和使用說明")
print("3. 🎯 狀態欄位有下拉選單：未開始、進行中、已完成")
print("4. 🎨 狀態會自動顯示顏色：綠色(已完成)、黃色(進行中)、紅色(未開始)")
print("5. 📊 統計會根據狀態欄位自動計算")
print("\n使用方式：")
print("- 直接在Excel中修改「狀態」欄位")
print("- 填寫「完成日期」和「備註」")
print("- 統計數據會自動更新")
